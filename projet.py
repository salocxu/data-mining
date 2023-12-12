import csv
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#jfdjd

# faure en sorte de choper directmeent les colonnes qui nous interessent ou prendre sa position 
def supprimer_sous_lim(data : pd.DataFrame, limite : float, liste_colonnes : list ):
    asuppr = (data[liste_colonnes] <= limite).any(axis=1)
    data = (data[~asuppr])
    print(data.isna().sum())
    data = data.dropna()
    print(data.isna().sum())
    return data

'''
def trouver_alt(lat, lon):
    
    return alt
'''

copier = 'oui'

# On parcourt les lignes du CSV
n= 50000
# pas besoin de toujours le faire
    

#voir = data.duplicated(subset='timestamps_UTC', keep='first')
valeur0inutile = ['RS_E_InAirTemp_PC1','RS_E_InAirTemp_PC2', 'RS_E_OilPress_PC1','RS_E_OilPress_PC2','RS_E_RPM_PC1','RS_E_RPM_PC2','RS_E_WatTemp_PC1','RS_E_WatTemp_PC2','RS_T_OilTemp_PC1','RS_T_OilTemp_PC2' ]


if copier == 'oui':
    data = pd.read_csv('../ar41_for_ulb.csv', sep=';',nrows=n)
    data = supprimer_sous_lim(data, 0, valeur0inutile )
    
    data[['date', 'hour']] = data['timestamps_UTC'].str.split(expand=True) 
    data.to_excel('data_tabmoins.xlsx', index=False)

#travailler à partir du excel

''' 
data = pd.read_excel('data_tabmoins.xlsx')
# je souhaite que tu me fasses une moyenne des donnees sur les colonnes 3 à fin

avg = data.iloc[:, 5:].mean(axis=0)
workbook = Workbook()
sheet = workbook.active

for index in range (len(avg)):
    sheet.cell(row=1, column=index + 1, value= data.columns[index + 5])
    sheet.cell(row=2, column=index + 1, value= avg[index])
    # modifier param cases
    column_letter = get_column_letter(index + 1)
    sheet.column_dimensions[column_letter].width = 20


workbook.save('data_avgmoins.xlsx')
workbook.close()
'''